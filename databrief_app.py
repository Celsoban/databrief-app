import streamlit as st
import pandas as pd
import anthropic
import json
import io

# ── Página ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DataBrief",
    page_icon="✦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── CSS — fiel ao React ───────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=IBM+Plex+Mono:wght@400;500&family=Inter:wght@300;400;500&display=swap');

*, *::before, *::after { box-sizing: border-box; }

html, body,
[data-testid="stAppViewContainer"],
[data-testid="stApp"],
[data-testid="stVerticalBlock"],
.stApp {
    background-color: #0A0C10 !important;
    color: #E8EAF0 !important;
    font-family: 'Inter', sans-serif !important;
}

#MainMenu, header[data-testid="stHeader"], footer { display: none !important; }

.block-container {
    padding: 28px 32px !important;
    max-width: 1100px !important;
    margin: 0 auto !important;
    background: transparent !important;
}

::-webkit-scrollbar { width: 4px; }
::-webkit-scrollbar-track { background: #111318; }
::-webkit-scrollbar-thumb { background: #1E2230; border-radius: 2px; }

@keyframes fadeUp {
    from { opacity: 0; transform: translateY(16px); }
    to   { opacity: 1; transform: translateY(0); }
}
@keyframes shimmer {
    0%   { background-position: -200% center; }
    100% { background-position:  200% center; }
}

/* ── Logo ── */
.db-logo { font-family:'Syne',sans-serif; font-weight:800; font-size:24px; color:#E8EAF0; letter-spacing:-.01em; line-height:1; }
.db-logo span { color:#00E5C0; }
.db-subtitle { font-family:'IBM Plex Mono',monospace; font-size:10px; color:#6B7280; letter-spacing:.12em; text-transform:uppercase; margin-top:3px; }
.db-logo-icon {
    width:34px; height:34px;
    background: linear-gradient(135deg,#00E5C0,#0099FF);
    border-radius:9px;
    display:inline-flex; align-items:center; justify-content:center;
    font-size:16px; margin-right:10px; vertical-align:middle;
}

/* ── Tags ── */
.db-tag {
    display:inline-flex; align-items:center; gap:5px;
    background:#00E5C020; color:#00E5C0;
    border:1px solid #00E5C030; border-radius:20px;
    padding:3px 11px; font-size:11px;
    font-family:'IBM Plex Mono',monospace; letter-spacing:.06em; font-weight:500;
    margin-right:6px;
}

/* ── Section title ── */
.db-section-title {
    font-family:'Syne',sans-serif; font-weight:700; font-size:11px;
    letter-spacing:.12em; text-transform:uppercase;
    color:#6B7280; margin-bottom:14px; margin-top:0;
}

/* ── KPI Card ── */
.db-kpi {
    background:#111318; border:1px solid #1E2230; border-radius:14px;
    padding:20px 24px; animation:fadeUp .5s ease both;
    transition: border-color .2s, transform .2s;
}
.db-kpi:hover { border-color:#00E5C050; transform:translateY(-2px); }
.db-kpi-label { font-family:'IBM Plex Mono',monospace; font-size:11px; color:#6B7280; letter-spacing:.08em; text-transform:uppercase; margin-bottom:8px; }
.db-kpi-value { font-family:'Syne',sans-serif; font-weight:700; font-size:28px; color:#00E5C0; line-height:1; }
.db-kpi-sub { font-size:12px; color:#6B7280; margin-top:6px; line-height:1.4; }

/* ── Insight Card ── */
.db-insight-card {
    background:#111318; border:1px solid #1E2230; border-radius:14px;
    padding:22px 26px; animation:fadeUp .5s ease both;
}
.db-insight-line {
    border-left:2px solid #00E5C0; padding-left:14px;
    margin:10px 0; font-size:14px; color:#E8EAF0; line-height:1.65;
}

/* ── Chat bubbles ── */
[data-testid="stChatMessage"] { background:transparent !important; border:none !important; }
[data-testid="stChatMessageContent"] { font-family:'Inter',sans-serif !important; font-size:14px !important; line-height:1.7 !important; color:#E8EAF0 !important; }

/* ── Chat input ── */
[data-testid="stChatInput"] { background:#111318 !important; border:1px solid #1E2230 !important; border-radius:10px !important; }
[data-testid="stChatInput"]:focus-within { border-color:#00E5C060 !important; box-shadow:none !important; }
[data-testid="stChatInput"] textarea { background:transparent !important; color:#E8EAF0 !important; font-family:'Inter',sans-serif !important; font-size:14px !important; }
[data-testid="stChatInput"] textarea::placeholder { color:#6B7280 !important; }
[data-testid="stChatInput"] button { background:#00E5C0 !important; color:#000 !important; border-radius:8px !important; }

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    background:#111318 !important; border:2px dashed #1E2230 !important;
    border-radius:14px !important; padding:24px !important;
    transition:border-color .25s, background .25s !important;
}
[data-testid="stFileUploader"]:hover { border-color:#00E5C0 !important; background:#00E5C008 !important; }
[data-testid="stFileUploader"] label { color:#E8EAF0 !important; font-family:'Syne',sans-serif !important; font-size:15px !important; font-weight:700 !important; }
[data-testid="stFileUploader"] small, [data-testid="stFileUploaderDropzoneInstructions"] { color:#6B7280 !important; font-size:13px !important; }
[data-testid="stFileUploader"] button { background:#00E5C0 !important; color:#000 !important; font-family:'Syne',sans-serif !important; font-weight:700 !important; border:none !important; border-radius:8px !important; }

/* ── Botão ghost ── */
.stButton > button {
    background:transparent !important; color:#6B7280 !important;
    border:1px solid #1E2230 !important; border-radius:8px !important;
    font-family:'Inter',sans-serif !important; font-size:13px !important;
    padding:8px 18px !important; transition:color .2s, border-color .2s !important;
}
.stButton > button:hover { color:#E8EAF0 !important; border-color:#6B7280 !important; background:transparent !important; }

/* ── Divider ── */
hr { border-color:#1E2230 !important; margin:0 0 28px 0 !important; }

/* ── Shimmer skeleton ── */
.db-shimmer {
    background: linear-gradient(90deg,#111318 25%,#1E2230 50%,#111318 75%);
    background-size:200% 100%; animation:shimmer 1.5s infinite;
    border-radius:14px; height:96px;
}

/* ── Glow fundo upload ── */
.db-glow {
    position:fixed; top:-20%; left:50%; transform:translateX(-50%);
    width:80vw; height:50vh;
    background:radial-gradient(ellipse,#00E5C012 0%,transparent 70%);
    pointer-events:none; z-index:0;
}

/* ── Upload center ── */
.db-upload-title { font-family:'Syne',sans-serif; font-weight:800; font-size:34px; color:#E8EAF0; letter-spacing:-.02em; line-height:1.2; margin-bottom:12px; }
.db-upload-title span { color:#00E5C0; }
.db-upload-sub { color:#6B7280; font-size:15px; line-height:1.7; max-width:420px; margin:0 auto 32px auto; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
def read_excel_clean(file) -> pd.DataFrame:
    """Lê Excel (.xlsx e .xls) ignorando fórmulas, usando só a aba com mais dados."""
    import io
    raw = file.read() if hasattr(file, "read") else open(file, "rb").read()
    fname = getattr(file, "name", "") or ""
    ext = fname.split(".")[-1].lower() if "." in fname else ""

    # .xls (formato legado) — usa xlrd
    if ext == "xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=raw)
            # Escolhe a aba com mais dados
            best = max(wb.sheets(), key=lambda s: s.nrows * s.ncols)

            # Detecta a linha de cabeçalho real:
            # procura a primeira linha com pelo menos 3 células não-vazias
            header_idx = 0
            for r in range(min(20, best.nrows)):
                filled = sum(1 for c in range(best.ncols)
                             if str(best.cell_value(r, c)).strip() not in ("", "None", "nan"))
                if filled >= 3:
                    header_idx = r
                    break

            headers = []
            for c in range(best.ncols):
                v = str(best.cell_value(header_idx, c)).strip()
                headers.append(v if v and v not in ("None","nan") else f"Col{c}")

            rows = []
            for r in range(header_idx + 1, best.nrows):
                rows.append([best.cell_value(r, c) for c in range(best.ncols)])
            df = pd.DataFrame(rows, columns=headers)
            print(f"[DataBrief] XLS lido — header na linha {header_idx}, {len(df)} linhas", flush=True)
            # Log valores brutos para diagnóstico
            if best.nrows > header_idx + 2:
                sample_row = [best.cell_value(header_idx + 2, c) for c in range(min(6, best.ncols))]
                sample_types = [best.cell_type(header_idx + 2, c) for c in range(min(6, best.ncols))]
                print(f"[DataBrief] XLS amostra linha {header_idx+2}: valores={sample_row} tipos={sample_types}", flush=True)
        except ImportError:
            df = pd.read_excel(io.BytesIO(raw), engine="xlrd")
    else:
        # .xlsx — usa openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        best_sheet = max(wb.sheetnames, key=lambda s: wb[s].max_row * wb[s].max_column)
        ws = wb[best_sheet]
        data = list(ws.values)
        if not data:
            raise ValueError("Planilha vazia.")
        headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(data[0])]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=headers)

    # Limpeza comum
    df = df.loc[:, df.apply(lambda col: not col.astype(str).str.startswith("=").all())]
    df = df.dropna(how="all").reset_index(drop=True)

    # Converte colunas que parecem numéricas mas vieram como texto
    # (comum em DREs com formatação de moeda)
    import re as _re
    def _try_numeric(col):
        def _parse(v):
            if v is None: return v
            # Se já é número, retorna direto — não processar novamente
            if isinstance(v, (int, float)):
                import math
                return None if math.isnan(v) else v
            s = str(v).strip()
            if s in ("", "None", "nan", "-", "0.0", "0"): return None
            # Só tenta converter se parece moeda/número texto
            # Remove R$, espaços — mas só remove ponto se há vírgula (formato BR)
            s = s.replace("R$","").replace(" ","").strip()
            if "," in s:
                # Formato BR: 1.234,56 → remover ponto de milhar, trocar vírgula
                s = s.replace(".","").replace(",",".")
            try: return float(s)
            except: return v
        converted = col.apply(_parse)
        non_null = converted.dropna()
        num_count = sum(1 for v in non_null if isinstance(v, (int, float)))
        if len(non_null) > 0 and num_count / len(non_null) >= 0.5:
            return pd.to_numeric(converted, errors="coerce")
        return col

    for c in df.columns:
        if df[c].dtype == object:
            df[c] = _try_numeric(df[c])

    num_after = len(df.select_dtypes(include="number").columns)
    print(f"[DataBrief] Arquivo lido: {len(df)} linhas, {len(df.columns)} colunas ({num_after} numéricas)", flush=True)
    return df


@st.cache_resource
def get_client():
    import os
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    print(f"[DataBrief] API key presente: {bool(key)} | primeiros chars: {key[:12] if key else 'VAZIA'}", flush=True)
    return anthropic.Anthropic(api_key=key)

def _safe_val(v):
    import math
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    try:
        import pandas as _pd
        if isinstance(v, _pd.Timestamp): return v.isoformat() if not _pd.isna(v) else None
    except Exception: pass
    if hasattr(v, "item"): return v.item()
    if hasattr(v, "isoformat"): return v.isoformat()
    return v

# ── Detector de formato ───────────────────────────────────────────────────────
DRE_KEYWORDS = {"receita","despesa","custo","lucro","caixa","vendas","cmv",
                "margem","resultado","faturamento","bruto","líquido","liquido",
                "operacional","financeiro","tributo","imposto","folha","salario",
                "salário","aluguel","total","subtotal","ebitda"}

def detect_format(df: pd.DataFrame) -> str:
    """Retorna 'dre' ou 'tabular' baseado na estrutura da planilha."""
    import unicodedata

    def normalize(s):
        """Remove acentos e converte para minúsculo."""
        return ''.join(c for c in unicodedata.normalize('NFD', str(s).lower())
                       if unicodedata.category(c) != 'Mn')

    DRE_KEYWORDS_NORM = {normalize(k) for k in DRE_KEYWORDS}

    num_cols = df.select_dtypes(include="number").columns
    num_ratio = len(num_cols) / max(len(df.columns), 1)

    # Log para diagnóstico
    first_col_vals = df.iloc[:, 0].astype(str).tolist()[:10] if len(df.columns) > 0 else []
    print(f"[DataBrief] detect_format: {len(df)} linhas, {len(df.columns)} colunas, "
          f"{len(num_cols)} numéricas, ratio={num_ratio:.2f}", flush=True)
    print(f"[DataBrief] primeiros valores col0: {first_col_vals[:5]}", flush=True)

    # Sinal 1: muitas colunas numéricas relativas ao total (típico de DRE com dias)
    if num_ratio > 0.5 and len(df) < 300:
        # Sinal 2: primeiras colunas têm palavras-chave contábeis (normalizado)
        for col_idx in range(min(3, len(df.columns))):
            col_vals = df.iloc[:, col_idx].astype(str).apply(normalize)
            matches = col_vals.apply(
                lambda v: any(kw in v for kw in DRE_KEYWORDS_NORM)
            ).sum()
            print(f"[DataBrief] col{col_idx} matches: {matches}", flush=True)
            if matches >= 2:
                print(f"[DataBrief] Formato detectado: DRE ({matches} palavras-chave)", flush=True)
                return "dre"

    # Sinal alternativo: menos de 5 colunas texto e mais de 20 colunas numéricas
    text_cols = len(df.columns) - len(num_cols)
    if text_cols <= 3 and len(num_cols) >= 20 and len(df) < 300:
        print(f"[DataBrief] Formato detectado: DRE (estrutura numérica densa)", flush=True)
        return "dre"

    print(f"[DataBrief] Formato detectado: tabular", flush=True)
    return "tabular"

def _safe_float(v):
    import math
    try:
        f = float(v)
        return None if math.isnan(f) or math.isinf(f) else round(f, 2)
    except Exception:
        return None

def build_summary_dre(df: pd.DataFrame) -> str:
    """Summary para planilhas DRE: linhas=contas, colunas=períodos."""
    import math

    # Identifica linha de cabeçalho dos períodos (procura linha com mais números)
    header_row = 0
    max_nums = 0
    for i in range(min(15, len(df))):
        nums = sum(1 for v in df.iloc[i] if _safe_float(v) is not None or str(v).strip().isdigit())
        if nums > max_nums:
            max_nums = nums
            header_row = i

    # Usa a linha identificada como cabeçalho dos períodos
    period_headers = [str(df.iloc[header_row, c]).strip() for c in range(len(df.columns))]

    # Identifica colunas de categoria (texto) e de valor (número)
    # Geralmente as 2 primeiras colunas são categoria/subcategoria
    cat_cols = []
    val_col_indices = []
    for c in range(len(df.columns)):
        col_vals = df.iloc[header_row+1:, c]
        num_count = sum(1 for v in col_vals if _safe_float(v) is not None)
        if num_count > len(col_vals) * 0.5:
            val_col_indices.append(c)
        else:
            cat_cols.append(c)

    # Processa contas — ignora linhas de subtotal (células mescladas ou vazias na categoria)
    contas = {}
    subtotal_keywords = {"total","subtotal","resultado","líquido","liquido","bruto","ebitda"}

    for i in range(header_row + 1, len(df)):
        row = df.iloc[i]

        # Nome da conta: concatena colunas de categoria não vazias
        nome_parts = []
        for c in cat_cols[:2]:
            v = str(row.iloc[c]).strip() if row.iloc[c] is not None else ""
            if v and v.lower() not in ("none","nan",""):
                nome_parts.append(v)
        if not nome_parts:
            continue
        nome = " > ".join(nome_parts)

        # É subtotal?
        is_subtotal = any(kw in nome.lower() for kw in subtotal_keywords)

        # Coleta valores por período — pega só a coluna "Total" se existir
        valores = {}
        total_col = None
        for c in val_col_indices:
            header = period_headers[c].lower()
            if "total" in header or header in ("total","t"):
                total_col = c
                break

        if total_col is not None:
            v = _safe_float(row.iloc[total_col])
            if v is not None:
                valores["total"] = v
        else:
            # Sem coluna total — soma todos os valores numéricos da linha
            vals = [_safe_float(row.iloc[c]) for c in val_col_indices if _safe_float(row.iloc[c]) is not None]
            if vals:
                valores["total"] = round(sum(vals), 2)
                # Guarda também min/max dos períodos para tendência
                valores["periodo_min"] = min(vals)
                valores["periodo_max"] = max(vals)
                valores["media_diaria"] = round(sum(vals)/len(vals), 2)

        if valores:
            contas[nome] = {"subtotal": is_subtotal, **valores}

    # Separa subtotais das contas detalhadas
    subtotais = {k: v for k, v in contas.items() if v["subtotal"]}
    detalhes  = {k: v for k, v in contas.items() if not v["subtotal"]}

    # Top contas por valor total
    top_contas = sorted(detalhes.items(), key=lambda x: abs(x[1].get("total", 0)), reverse=True)[:10]

    # Log para diagnóstico dos valores
    print(f"[DataBrief] DRE: {len(subtotais)} subtotais, {len(detalhes)} detalhes", flush=True)
    print(f"[DataBrief] DRE subtotais: {list(subtotais.items())[:3]}", flush=True)
    print(f"[DataBrief] DRE top3 contas: {top_contas[:3]}", flush=True)

    payload = {
        "ATENCAO": "Planilha DRE — linhas são contas contábeis, colunas são períodos. Use APENAS os dados abaixo.",
        "formato": "DRE",
        "total_contas": len(detalhes),
        "periodos_identificados": len(val_col_indices),
        "subtotais": {k: v["total"] for k, v in subtotais.items() if "total" in v},
        "top_contas_por_valor": {k: v for k, v in top_contas},
    }
    result = json.dumps(payload, ensure_ascii=False, separators=(',',':'), default=str)
    print(f"[DataBrief] summary DRE — {len(result)} chars, {len(detalhes)} contas", flush=True)
    return result

def build_summary_tabular(df: pd.DataFrame) -> str:
    """Summary para planilhas tabulares convencionais (cada linha = 1 registro)."""
    import math, pandas as _pd

    # Limita colunas (nunca linhas — agregados usam o dataset completo)
    original_cols = len(df.columns)
    if len(df.columns) > 30:
        num = df.select_dtypes(include="number").columns.tolist()[:15]
        cat = []
        for c in df.columns:
            if c not in num:
                try:
                    if int(df[c].nunique()) <= 20:
                        cat.append(c)
                except Exception:
                    pass
        cat = cat[:15]
        df = df[num + cat]
    print(f"[DataBrief] {len(df)} linhas | colunas: {original_cols} → {len(df.columns)}", flush=True)

    # Converte datas para string
    for col in df.columns:
        if _pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d").fillna("")
        else:
            def _to_safe(v):
                if v is None: return None
                if isinstance(v, float) and math.isnan(v): return None
                if isinstance(v, _pd.Timestamp): return v.strftime("%Y-%m-%d") if not _pd.isna(v) else None
                try:
                    if _pd.isna(v): return None
                except Exception: pass
                if hasattr(v, "item"): return v.item()
                return v
            df[col] = df[col].apply(_to_safe)

    num_cols  = df.select_dtypes(include="number").columns.tolist()
    text_cols = [c for c in df.columns if c not in num_cols]

    agg = {}
    for col in num_cols:
        vals = df[col].dropna()
        if len(vals) == 0: continue
        agg[col] = {
            "soma":      round(float(vals.sum()),  2),
            "media":     round(float(vals.mean()), 2),
            "min":       round(float(vals.min()),  2),
            "max":       round(float(vals.max()),  2),
            "nao_nulos": int(vals.count()),
        }

    freq = {}
    for col in text_cols:
        unique = df[col].nunique(dropna=True)
        if unique <= 20:
            f = df[col].value_counts(dropna=True).head(5).to_dict()
            freq[col] = {str(k): int(v) for k, v in f.items()}

    cols_amostra = num_cols[:8] + [c for c in text_cols if c in freq][:5]
    amostra = df[cols_amostra].head(3).where(df[cols_amostra].head(3).notna(), other=None).to_dict(orient="records")

    payload = {
        "ATENCAO": "Use APENAS os dados abaixo. Nunca invente ou estime valores.",
        "formato": "tabular",
        "total_registros_exato": len(df),
        "colunas_numericas": num_cols,
        "colunas_categoricas": list(freq.keys()),
        "agregados_numericos": agg,
        "frequencias_categoricas": freq,
        "amostra_3_linhas": amostra,
    }
    result = json.dumps(payload, ensure_ascii=False, separators=(',',':'), default=str)
    print(f"[DataBrief] summary tabular — {len(result)} chars", flush=True)
    return result

def build_summary(df: pd.DataFrame) -> str:
    """Detecta formato e delega para o summary correto."""
    fmt = detect_format(df)
    if fmt == "dre":
        return build_summary_dre(df)
    return build_summary_tabular(df)

def call_claude(system: str, message: str, max_tok: int = 800) -> str:
    import time, httpx, os
    t0 = time.time()
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    print(f"[DataBrief] Chamando API Claude... max_tokens={max_tok} | key={key[:12] if key else 'VAZIA'}", flush=True)
    try:
        client = anthropic.Anthropic(
            api_key=key,
            timeout=httpx.Timeout(45.0, connect=10.0)
        )
        r = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=max_tok,
            system=system,
            messages=[{"role": "user", "content": message}],
        )
        print(f"[DataBrief] API respondeu em {time.time()-t0:.1f}s", flush=True)
        return r.content[0].text
    except Exception as e:
        print(f"[DataBrief] ERRO na API: {type(e).__name__}: {e}", flush=True)
        raise

def analyze_data(df: pd.DataFrame, tipo_negocio: str = "") -> dict:
    print(f"[DataBrief] analyze_data iniciado — {len(df)} linhas", flush=True)
    summary  = build_summary(df)
    print(f"[DataBrief] summary gerado — {len(summary)} chars", flush=True)
    contexto = f"Tipo de negócio: {tipo_negocio}." if tipo_negocio else "Tipo de negócio: não informado — identifique pelo contexto dos dados."

    # Detecta se é DRE para ajustar o prompt
    is_dre = '"formato":"DRE"' in summary or '"formato": "DRE"' in summary

    if is_dre:
        system = (
            "Você é o DataBrief, analista financeiro especialista em DRE para pequenos negócios.\n"
            + contexto + "\n"
            + "DRE recebido:\n" + summary + "\n"
            + "Responda em português, de forma clara e direta para um dono de negócio sem conhecimento técnico."
        )
        prompt = """Analise o DRE e retorne SOMENTE JSON válido (sem markdown, sem texto extra):
{
  "kpis": [
    {"label":"...","value":"...","sub":"..."}
  ],
  "melhorou": ["ponto positivo financeiro 1", "ponto positivo financeiro 2"],
  "piorou":   ["ponto negativo financeiro 1", "ponto negativo financeiro 2"],
  "alertas":  ["alerta financeiro 1", "alerta financeiro 2"],
  "recomendacao": "Uma ação financeira concreta e imediata que o dono deve tomar agora."
}
REGRAS CRÍTICAS:
- Use EXCLUSIVAMENTE os valores de "subtotais" e "top_contas_por_valor"
- KPIs devem mostrar: receita total, maior despesa, margem ou resultado se disponível
- "melhorou": contas com valores positivos ou receitas expressivas
- "piorou": contas com maiores despesas ou valores negativos
- "alertas": despesas elevadas, contas sem valor, desequilíbrios receita/despesa
- "recomendacao": 1 ação prática baseada nos números do DRE
- NUNCA invente valores. Se não houver dados suficientes, retorne lista vazia []"""
    else:
        system = (
            "Você é o DataBrief, analista de dados especialista em pequenos negócios.\n"
            + contexto + "\n"
            + "Planilha recebida:\n" + summary + "\n"
            + "Responda em português, de forma clara e direta para um dono de negócio sem conhecimento técnico."
        )
        prompt = """Analise os dados e retorne SOMENTE JSON válido (sem markdown, sem texto extra):
{
  "kpis": [
    {"label":"...","value":"...","sub":"..."}
  ],
  "melhorou": ["ponto positivo 1", "ponto positivo 2"],
  "piorou":   ["ponto negativo 1", "ponto negativo 2"],
  "alertas":  ["anomalia ou atenção 1", "anomalia ou atenção 2"],
  "recomendacao": "Uma ação concreta e imediata que o dono deve tomar agora."
}
REGRAS CRÍTICAS:
- Use EXCLUSIVAMENTE os números dos campos "agregados_numericos" e "frequencias_categoricas"
- O total de registros é EXATAMENTE o campo "total_registros_exato" — nunca use outro número
- NUNCA invente ou estime valores que não estejam explicitamente nos dados
- Valores dos KPIs devem vir diretamente dos dados, com unidade (ex: "R$ 2.344.793", "71 registros")
- "melhorou" e "piorou": comparações reais dentro dos dados (top vs bottom de frequências)
- "alertas": concentrações, valores extremos, campos com muitos nulos
- "recomendacao": 1 única ação prática baseada nos dados reais
- Se não houver dados suficientes para alguma seção, retorne lista vazia []"""

    raw = call_claude(system, prompt)
    return json.loads(raw.replace("```json","").replace("```","").strip())

# ── Session state ─────────────────────────────────────────────────────────────
for k, v in [("df",None),("file_name",""),("analysis",None),("messages",[]),("tipo_negocio","")]:
    if k not in st.session_state:
        st.session_state[k] = v

TIPOS_NEGOCIO = [
    {"emoji": "🍽️", "label": "Restaurante"},
    {"emoji": "🏪", "label": "Loja / Varejo"},
    {"emoji": "🏥", "label": "Clínica / Saúde"},
    {"emoji": "🏋️", "label": "Academia / Fitness"},
    {"emoji": "💼", "label": "Agência / Serviços"},
    {"emoji": "📦", "label": "Outro"},
]

# ═══════════════════════════════════════════════════════
# TELA DE UPLOAD
# ═══════════════════════════════════════════════════════
if st.session_state.df is None:

    st.markdown('<div class="db-glow"></div>', unsafe_allow_html=True)
    st.markdown("<br><br>", unsafe_allow_html=True)

    _, mid, _ = st.columns([1, 2, 1])
    with mid:
        # Logo
        st.markdown("""
        <div style="display:flex;align-items:center;justify-content:center;gap:10px;margin-bottom:36px;">
            <div class="db-logo-icon">✦</div>
            <div>
                <div class="db-logo">Data<span>Brief</span></div>
                <div class="db-subtitle">AI Data Analyst</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Headline
        st.markdown("""
        <div style="text-align:center;">
            <div class="db-upload-title">Seus dados.<br><span>Inteligência instantânea.</span></div>
            <div class="db-upload-sub">
                Faça upload da sua planilha e receba análise,
                KPIs e insights em segundos — com IA.
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Seleção de tipo de negócio ──
        st.markdown("""
        <div style="margin:24px 0 10px 0;">
            <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#6B7280;
                        letter-spacing:.12em;text-transform:uppercase;text-align:center;margin-bottom:12px;">
                Qual é o seu negócio?
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Botões de tipo — 3 colunas x 2 linhas
        row1 = TIPOS_NEGOCIO[:3]
        row2 = TIPOS_NEGOCIO[3:]
        for row in [row1, row2]:
            cols = st.columns(3)
            for col, tipo in zip(cols, row):
                with col:
                    selecionado = st.session_state.tipo_negocio == tipo["label"]
                    border = "#00E5C0" if selecionado else "#1E2230"
                    bg     = "#00E5C015" if selecionado else "#111318"
                    color  = "#00E5C0"  if selecionado else "#6B7280"
                    st.markdown(f"""
                    <div style="background:{bg};border:1px solid {border};border-radius:12px;
                                padding:12px 8px;text-align:center;cursor:pointer;
                                transition:all .2s;margin-bottom:8px;">
                        <div style="font-size:22px;margin-bottom:4px;">{tipo['emoji']}</div>
                        <div style="font-size:12px;font-family:'Syne',sans-serif;font-weight:700;color:{color};">
                            {tipo['label']}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    if st.button(tipo["label"], key=f"tipo_{tipo['label']}", use_container_width=True):
                        st.session_state.tipo_negocio = tipo["label"]
                        st.rerun()

        # Mostrar selecionado
        if st.session_state.tipo_negocio:
            tipo_sel = next(t for t in TIPOS_NEGOCIO if t["label"] == st.session_state.tipo_negocio)
            st.markdown(f"""
            <div style="text-align:center;margin:8px 0 20px 0;">
                <span class="db-tag">{tipo_sel['emoji']} {tipo_sel['label']} selecionado</span>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        # ── Uploader ──
        uploaded = st.file_uploader(
            "⬆  Arraste seu CSV ou Excel aqui, ou clique para selecionar",
            type=["csv", "xlsx", "xls"],
            label_visibility="visible",
        )

    if uploaded:
        try:
            ext = uploaded.name.split(".")[-1].lower()
            if ext in ("xlsx", "xls"):
                df = read_excel_clean(uploaded)
            else:
                df = pd.read_csv(uploaded, sep=None, engine="python")
            st.session_state.df        = df
            st.session_state.file_name = uploaded.name
            st.session_state.analysis  = None
            st.session_state.messages  = []
            st.rerun()
        except Exception as e:
            st.error(f"Erro ao ler o arquivo: {e}")

# ═══════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════
else:
    df = st.session_state.df

    # ── Header ────────────────────────────────────────────────────────────────
    c_logo, c_tags, c_btn = st.columns([2, 4, 1])
    with c_logo:
        st.markdown("""
        <div style="display:flex;align-items:center;gap:10px;padding-top:4px;">
            <div class="db-logo-icon">✦</div>
            <div>
                <div class="db-logo">Data<span>Brief</span></div>
                <div class="db-subtitle">AI Data Analyst</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with c_tags:
        st.markdown(f"""
        <div style="padding-top:10px;">
            <span class="db-tag">📄 {st.session_state.file_name}</span>
            <span class="db-tag">✦ {len(df)} linhas</span>
            <span class="db-tag">{len(df.columns)} colunas</span>
        </div>
        """, unsafe_allow_html=True)
    with c_btn:
        if st.button("Novo arquivo"):
            for k in ("df","file_name","analysis","messages"):
                st.session_state[k] = None if k in ("df","analysis") else ([] if k=="messages" else "")
            st.session_state.tipo_negocio = ""
            st.rerun()

    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Análise ────────────────────────────────────────────────────────────────
    if st.session_state.analysis is None:
        with st.spinner("✦  Analisando seus dados com IA…"):
            try:
                st.session_state.analysis = analyze_data(df, st.session_state.tipo_negocio)
            except Exception as e:
                st.session_state.analysis = {"kpis":[],"insights":[f"Erro: {e}"]}

    analysis = st.session_state.analysis

    # ── KPIs ──────────────────────────────────────────────────────────────────
    st.markdown('<p class="db-section-title">Visão Geral</p>', unsafe_allow_html=True)
    kpis = analysis.get("kpis", [])
    if kpis:
        cols = st.columns(len(kpis))
        for i, kpi in enumerate(kpis):
            with cols[i]:
                st.markdown(f"""
                <div class="db-kpi" style="animation-delay:{i*80}ms">
                    <div class="db-kpi-label">{kpi.get('label','')}</div>
                    <div class="db-kpi-value">{kpi.get('value','')}</div>
                    <div class="db-kpi-sub">{kpi.get('sub','')}</div>
                </div>
                """, unsafe_allow_html=True)
    else:
        for col in st.columns(3):
            with col:
                st.markdown('<div class="db-shimmer"></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Análise Estruturada ───────────────────────────────────────────────────
    st.markdown('<p class="db-section-title">Análise</p>', unsafe_allow_html=True)

    melhorou   = analysis.get("melhorou", [])
    piorou     = analysis.get("piorou", [])
    alertas    = analysis.get("alertas", [])
    recomendacao = analysis.get("recomendacao", "")

    col_a, col_b = st.columns(2)

    with col_a:
        if melhorou:
            items = "".join(f'<div class="db-insight-line" style="border-color:#00E5C0">✅ {i}</div>' for i in melhorou)
            st.markdown(f"""
            <div class="db-insight-card" style="margin-bottom:14px;">
                <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#00E5C0;letter-spacing:.1em;text-transform:uppercase;margin-bottom:10px;">O que melhorou</div>
                {items}
            </div>""", unsafe_allow_html=True)

        if alertas:
            items = "".join(f'<div class="db-insight-line" style="border-color:#FFB547">⚠️ {i}</div>' for i in alertas)
            st.markdown(f"""
            <div class="db-insight-card">
                <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#FFB547;letter-spacing:.1em;text-transform:uppercase;margin-bottom:10px;">Alertas</div>
                {items}
            </div>""", unsafe_allow_html=True)

    with col_b:
        if piorou:
            items = "".join(f'<div class="db-insight-line" style="border-color:#FF4D6D">📉 {i}</div>' for i in piorou)
            st.markdown(f"""
            <div class="db-insight-card" style="margin-bottom:14px;">
                <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#FF4D6D;letter-spacing:.1em;text-transform:uppercase;margin-bottom:10px;">O que piorou</div>
                {items}
            </div>""", unsafe_allow_html=True)

        if recomendacao:
            st.markdown(f"""
            <div class="db-insight-card" style="border-color:#00E5C040;background:linear-gradient(135deg,#00E5C008,#111318);">
                <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#00E5C0;letter-spacing:.1em;text-transform:uppercase;margin-bottom:10px;">★ Ação imediata</div>
                <div style="font-size:14px;color:#E8EAF0;line-height:1.65;font-weight:500;">{recomendacao}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Chat ──────────────────────────────────────────────────────────────────
    st.markdown('<p class="db-section-title">Converse com seus dados</p>', unsafe_allow_html=True)

    summary     = build_summary(df)
    tipo_ctx = f"Tipo de negócio: {st.session_state.tipo_negocio}. " if st.session_state.tipo_negocio else ""
    # Versão compacta do sumário para o chat — evita payload gigante a cada mensagem
    import json as _json
    _s = _json.loads(summary)
    summary_chat = _json.dumps({
        "total_registros": _s.get("total_registros_exato"),
        "agregados":       _s.get("agregados_numericos", {}),
        "frequencias":     {k: dict(list(v.items())[:5]) for k, v in _s.get("frequencias_categoricas", {}).items()},
    }, ensure_ascii=False)

    system_chat = f"""Você é o DataBrief, analista de dados especialista em pequenos negócios.
{tipo_ctx}Resumo dos dados da planilha:
{summary_chat}
Responda em português, de forma clara, objetiva e útil para tomada de decisão. Seja direto e conciso."""

    # Histórico do chat
    if not st.session_state.messages:
        st.markdown("""
        <div style="text-align:center;color:#6B7280;font-size:13px;padding:28px 0 8px 0;">
            💬  Faça qualquer pergunta sobre seus dados em linguagem natural.
        </div>
        """, unsafe_allow_html=True)

    for msg in st.session_state.messages:
        role_label = "Você" if msg["role"] == "user" else "DataBrief"
        bg = "#1E2A3A" if msg["role"] == "user" else "#12201A"
        border = "#2E4A6A" if msg["role"] == "user" else "#00E5C0"
        align = "flex-end" if msg["role"] == "user" else "flex-start"
        st.markdown(f"""
        <div style="display:flex;justify-content:{align};margin:8px 0;">
          <div style="background:{bg};border-left:3px solid {border};border-radius:8px;
                      padding:12px 16px;max-width:85%;">
            <div style="font-size:11px;color:#6B7280;margin-bottom:4px;">{role_label}</div>
            <div style="color:#E5E7EB;font-size:14px;">{msg["content"]}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

    # Input integrado ao layout
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    col_input, col_btn = st.columns([5, 1])
    with col_input:
        user_input = st.text_input(
            label="chat_input",
            label_visibility="collapsed",
            placeholder="Ex: Qual produto vendeu mais? Quem performa melhor?",
            key="chat_input_field"
        )
    with col_btn:
        send = st.button("Enviar →", use_container_width=True)

    if send and user_input:
        st.session_state.messages.append({"role":"user","content":user_input})
        with st.spinner("DataBrief está analisando…"):
            history = "\n".join(
                f"{'Usuário' if m['role']=='user' else 'DataBrief'}: {m['content']}"
                for m in st.session_state.messages
            )
            try:
                answer = call_claude(
                    system_chat,
                    f"Histórico:\n{history}\n\nResponda à última pergunta de forma clara e objetiva."
                )
            except Exception as e:
                answer = f"Erro ao conectar com a IA: {e}"
        st.session_state.messages.append({"role":"assistant","content":answer})
        st.rerun()
